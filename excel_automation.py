import pandas as pd
import numpy as np
import os
from datetime import datetime, date, timedelta
from typing import Dict, List, Optional, Tuple, Any
import streamlit as st
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import BarChart, Reference, LineChart
import json
import shutil

class ExcelAutomationManager:
    """
    Advanced Excel automation system for Smart Kids Attendance System
    Handles Excel file creation, formatting, data management, and reporting
    """
    
    def __init__(self):
        self.data_dir = "data"
        self.excel_dir = "excel_reports"
        self.templates_dir = "excel_templates"
        
        # File paths
        self.teachers_file = f"{self.data_dir}/teachers.xlsx"
        self.attendance_file = f"{self.data_dir}/attendance.xlsx"
        self.reports_file = f"{self.excel_dir}/attendance_reports.xlsx"
        self.summary_file = f"{self.excel_dir}/monthly_summary.xlsx"
        
        # Create directories
        self._create_directories()
        
        # Excel styling
        self.header_font = Font(bold=True, color="FFFFFF")
        self.header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        self.center_alignment = Alignment(horizontal='center', vertical='center')
        
    def _create_directories(self):
        """Create necessary directories"""
        for directory in [self.data_dir, self.excel_dir, self.templates_dir]:
            os.makedirs(directory, exist_ok=True)
    
    def create_excel_template(self, template_type: str) -> Tuple[bool, str]:
        """Create Excel templates for different purposes"""
        try:
            if template_type == "teachers":
                return self._create_teachers_template()
            elif template_type == "attendance":
                return self._create_attendance_template()
            elif template_type == "bulk_import":
                return self._create_bulk_import_template()
            elif template_type == "report":
                return self._create_report_template()
            else:
                return False, "Unknown template type"
        except Exception as e:
            return False, f"Error creating template: {str(e)}"
    
    def _create_teachers_template(self) -> Tuple[bool, str]:
        """Create teachers Excel template with formatting"""
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Teachers"
            
            # Headers
            headers = ['ID', 'Name', 'Department', 'Registration_Date', 'Face_Encoding_Path', 'Status', 'Email']
            ws.append(headers)
            
            # Format headers
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num)
                cell.font = self.header_font
                cell.fill = self.header_fill
                cell.border = self.border
                cell.alignment = self.center_alignment
            
            # Add sample data
            sample_data = [
                ['T001', 'John Doe', 'Mathematics', '2024-01-15', 'face_encodings/T001.pkl', 'Active', 'john.doe@school.com'],
                ['T002', 'Jane Smith', 'Science', '2024-01-16', 'face_encodings/T002.pkl', 'Active', 'jane.smith@school.com']
            ]
            
            for row_data in sample_data:
                ws.append(row_data)
            
            # Auto-adjust column widths
            self._auto_adjust_columns(ws)
            
            # Save template
            template_path = f"{self.templates_dir}/teachers_template.xlsx"
            wb.save(template_path)
            
            return True, f"Teachers template created at {template_path}"
            
        except Exception as e:
            return False, f"Error creating teachers template: {str(e)}"
    
    def _create_attendance_template(self) -> Tuple[bool, str]:
        """Create attendance Excel template"""
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Attendance"
            
            # Headers
            headers = ['Date', 'Teacher_ID', 'Name', 'Time_In', 'Status', 'Is_Holiday', 'Holiday_Name', 'Recognition_Confidence']
            ws.append(headers)
            
            # Format headers
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num)
                cell.font = self.header_font
                cell.fill = self.header_fill
                cell.border = self.border
                cell.alignment = self.center_alignment
            
            # Add sample data
            sample_data = [
                ['2024-01-15', 'T001', 'John Doe', '09:15:30', 'Present', False, '', 0.85],
                ['2024-01-15', 'T002', 'Jane Smith', '09:18:45', 'Present', False, '', 0.92]
            ]
            
            for row_data in sample_data:
                ws.append(row_data)
            
            # Auto-adjust column widths
            self._auto_adjust_columns(ws)
            
            # Save template
            template_path = f"{self.templates_dir}/attendance_template.xlsx"
            wb.save(template_path)
            
            return True, f"Attendance template created at {template_path}"
            
        except Exception as e:
            return False, f"Error creating attendance template: {str(e)}"
    
    def _create_bulk_import_template(self) -> Tuple[bool, str]:
        """Create bulk import template for teachers"""
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Bulk_Import_Teachers"
            
            # Headers
            headers = ['ID', 'Name', 'Department', 'Email', 'Notes']
            ws.append(headers)
            
            # Format headers
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num)
                cell.font = self.header_font
                cell.fill = self.header_fill
                cell.border = self.border
                cell.alignment = self.center_alignment
            
            # Add instructions
            instructions = [
                ['Instructions:', '', '', '', ''],
                ['1. Fill in teacher details in the rows below', '', '', '', ''],
                ['2. ID should be unique (e.g., T001, T002)', '', '', '', ''],
                ['3. Department examples: Mathematics, Science, English', '', '', '', ''],
                ['4. Email should be valid format', '', '', '', ''],
                ['5. Notes are optional', '', '', '', ''],
                ['', '', '', '', ''],
                ['Sample Data:', '', '', '', ''],
            ]
            
            for instruction in instructions:
                ws.append(instruction)
            
            # Add sample data
            sample_data = [
                ['T001', 'John Doe', 'Mathematics', 'john.doe@school.com', 'Head of Math Department'],
                ['T002', 'Jane Smith', 'Science', 'jane.smith@school.com', 'Physics Teacher'],
                ['T003', 'Mike Johnson', 'English', 'mike.johnson@school.com', 'Literature Specialist']
            ]
            
            for row_data in sample_data:
                ws.append(row_data)
            
            # Auto-adjust column widths
            self._auto_adjust_columns(ws)
            
            # Save template
            template_path = f"{self.templates_dir}/bulk_import_template.xlsx"
            wb.save(template_path)
            
            return True, f"Bulk import template created at {template_path}"
            
        except Exception as e:
            return False, f"Error creating bulk import template: {str(e)}"
    
    def _create_report_template(self) -> Tuple[bool, str]:
        """Create comprehensive report template"""
        try:
            wb = Workbook()
            
            # Summary Sheet
            ws_summary = wb.active
            ws_summary.title = "Summary"
            
            # Daily Attendance Sheet
            ws_daily = wb.create_sheet("Daily_Attendance")
            
            # Monthly Report Sheet
            ws_monthly = wb.create_sheet("Monthly_Report")
            
            # Teacher Statistics Sheet
            ws_stats = wb.create_sheet("Teacher_Statistics")
            
            # Format each sheet
            self._format_summary_sheet(ws_summary)
            self._format_daily_sheet(ws_daily)
            self._format_monthly_sheet(ws_monthly)
            self._format_stats_sheet(ws_stats)
            
            # Save template
            template_path = f"{self.templates_dir}/report_template.xlsx"
            wb.save(template_path)
            
            return True, f"Report template created at {template_path}"
            
        except Exception as e:
            return False, f"Error creating report template: {str(e)}"
    
    def import_teachers_from_excel(self, file_path: str) -> Tuple[bool, str, List[Dict]]:
        """Import teachers from Excel file"""
        try:
            # Read Excel file
            df = pd.read_excel(file_path)
            
            # Validate required columns
            required_columns = ['ID', 'Name', 'Department', 'Email']
            missing_columns = [col for col in required_columns if col not in df.columns]
            
            if missing_columns:
                return False, f"Missing required columns: {missing_columns}", []
            
            # Process each teacher
            imported_teachers = []
            errors = []
            
            for index, row in df.iterrows():
                try:
                    teacher_data = {
                        'ID': str(row['ID']).strip(),
                        'Name': str(row['Name']).strip(),
                        'Department': str(row['Department']).strip(),
                        'Email': str(row['Email']).strip(),
                        'Registration_Date': datetime.now().strftime('%Y-%m-%d'),
                        'Status': 'Active',
                        'Face_Encoding_Path': f"face_encodings/{row['ID']}.pkl"
                    }
                    
                    # Validate data
                    if not teacher_data['ID'] or not teacher_data['Name']:
                        errors.append(f"Row {index + 2}: ID and Name are required")
                        continue
                    
                    imported_teachers.append(teacher_data)
                    
                except Exception as e:
                    errors.append(f"Row {index + 2}: {str(e)}")
            
            if errors:
                return False, f"Import errors: {'; '.join(errors)}", imported_teachers
            
            return True, f"Successfully processed {len(imported_teachers)} teachers", imported_teachers
            
        except Exception as e:
            return False, f"Error importing from Excel: {str(e)}", []
    
    def export_attendance_report(self, start_date: str, end_date: str, 
                                teacher_ids: List[str] = None) -> Tuple[bool, str]:
        """Export comprehensive attendance report"""
        try:
            # Load attendance data
            attendance_df = pd.read_excel(self.attendance_file)
            
            # Filter by date range
            mask = (attendance_df['Date'] >= start_date) & (attendance_df['Date'] <= end_date)
            filtered_df = attendance_df[mask]
            
            # Filter by teacher IDs if specified
            if teacher_ids:
                filtered_df = filtered_df[filtered_df['Teacher_ID'].isin(teacher_ids)]
            
            # Create workbook
            wb = Workbook()
            
            # Raw Data Sheet
            ws_raw = wb.active
            ws_raw.title = "Raw_Data"
            
            # Add data to raw sheet
            for r in dataframe_to_rows(filtered_df, index=False, header=True):
                ws_raw.append(r)
            
            # Format raw data sheet
            self._format_data_sheet(ws_raw, filtered_df.columns.tolist())
            
            # Summary Sheet
            ws_summary = wb.create_sheet("Summary")
            self._create_attendance_summary(ws_summary, filtered_df, start_date, end_date)
            
            # Daily Statistics Sheet
            ws_daily_stats = wb.create_sheet("Daily_Statistics")
            self._create_daily_statistics(ws_daily_stats, filtered_df)
            
            # Teacher Performance Sheet
            ws_performance = wb.create_sheet("Teacher_Performance")
            self._create_teacher_performance(ws_performance, filtered_df)
            
            # Save report
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            report_path = f"{self.excel_dir}/attendance_report_{timestamp}.xlsx"
            wb.save(report_path)
            
            return True, f"Report exported to {report_path}"
            
        except Exception as e:
            return False, f"Error exporting report: {str(e)}"
    
    def create_monthly_summary(self, year: int, month: int) -> Tuple[bool, str]:
        """Create monthly attendance summary"""
        try:
            # Load data
            attendance_df = pd.read_excel(self.attendance_file)
            teachers_df = pd.read_excel(self.teachers_file)
            
            # Filter for the month
            month_str = f"{year}-{month:02d}"
            monthly_data = attendance_df[attendance_df['Date'].str.startswith(month_str)]
            
            # Create workbook
            wb = Workbook()
            ws = wb.active
            ws.title = f"Summary_{year}_{month:02d}"
            
            # Title
            ws.merge_cells('A1:H1')
            title_cell = ws['A1']
            title_cell.value = f"Monthly Attendance Summary - {datetime(year, month, 1).strftime('%B %Y')}"
            title_cell.font = Font(size=16, bold=True)
            title_cell.alignment = self.center_alignment
            
            # Headers
            headers = ['Teacher ID', 'Name', 'Department', 'Days Present', 'Total Days', 'Attendance %', 'Avg Time', 'Status']
            ws.append([''])  # Empty row
            ws.append(headers)
            
            # Format headers
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=3, column=col_num)
                cell.font = self.header_font
                cell.fill = self.header_fill
                cell.border = self.border
                cell.alignment = self.center_alignment
            
            # Calculate statistics for each teacher
            for _, teacher in teachers_df.iterrows():
                if teacher['Status'] != 'Active':
                    continue
                
                teacher_attendance = monthly_data[monthly_data['Teacher_ID'] == teacher['ID']]
                days_present = len(teacher_attendance)
                
                # Calculate working days in month (excluding weekends)
                total_days = self._get_working_days_in_month(year, month)
                
                attendance_percentage = (days_present / total_days * 100) if total_days > 0 else 0
                
                # Calculate average time
                if not teacher_attendance.empty:
                    avg_time = self._calculate_average_time(teacher_attendance['Time_In'].tolist())
                else:
                    avg_time = "N/A"
                
                # Status based on attendance
                if attendance_percentage >= 90:
                    status = "Excellent"
                elif attendance_percentage >= 80:
                    status = "Good"
                elif attendance_percentage >= 70:
                    status = "Average"
                else:
                    status = "Poor"
                
                row_data = [
                    teacher['ID'],
                    teacher['Name'],
                    teacher['Department'],
                    days_present,
                    total_days,
                    f"{attendance_percentage:.1f}%",
                    avg_time,
                    status
                ]
                
                ws.append(row_data)
            
            # Auto-adjust columns
            self._auto_adjust_columns(ws)
            
            # Add chart
            self._add_attendance_chart(ws, len(teachers_df))
            
            # Save file
            summary_path = f"{self.excel_dir}/monthly_summary_{year}_{month:02d}.xlsx"
            wb.save(summary_path)
            
            return True, f"Monthly summary created at {summary_path}"
            
        except Exception as e:
            return False, f"Error creating monthly summary: {str(e)}"
    
    def backup_excel_files(self) -> Tuple[bool, str]:
        """Create backup of all Excel files"""
        try:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            backup_dir = f"data/backups/excel_backup_{timestamp}"
            os.makedirs(backup_dir, exist_ok=True)
            
            # Files to backup
            files_to_backup = [
                self.teachers_file,
                self.attendance_file,
                f"{self.excel_dir}/*.xlsx"
            ]
            
            backup_count = 0
            for file_pattern in files_to_backup:
                if '*' in file_pattern:
                    # Handle wildcard patterns
                    import glob
                    for file_path in glob.glob(file_pattern):
                        if os.path.exists(file_path):
                            shutil.copy2(file_path, backup_dir)
                            backup_count += 1
                else:
                    if os.path.exists(file_pattern):
                        shutil.copy2(file_pattern, backup_dir)
                        backup_count += 1
            
            return True, f"Backed up {backup_count} files to {backup_dir}"
            
        except Exception as e:
            return False, f"Error creating backup: {str(e)}"
    
    def validate_excel_data(self, file_path: str, data_type: str) -> Tuple[bool, str, List[str]]:
        """Validate Excel data before import"""
        try:
            df = pd.read_excel(file_path)
            errors = []
            warnings = []
            
            if data_type == "teachers":
                # Check required columns
                required_cols = ['ID', 'Name', 'Department', 'Email']
                missing_cols = [col for col in required_cols if col not in df.columns]
                if missing_cols:
                    errors.append(f"Missing columns: {missing_cols}")
                
                # Validate data
                for index, row in df.iterrows():
                    row_num = index + 2  # Excel row number
                    
                    # Check ID format
                    if pd.isna(row.get('ID')) or str(row['ID']).strip() == '':
                        errors.append(f"Row {row_num}: ID is required")
                    
                    # Check name
                    if pd.isna(row.get('Name')) or str(row['Name']).strip() == '':
                        errors.append(f"Row {row_num}: Name is required")
                    
                    # Check email format
                    email = str(row.get('Email', '')).strip()
                    if email and '@' not in email:
                        warnings.append(f"Row {row_num}: Invalid email format")
            
            elif data_type == "attendance":
                # Validate attendance data
                required_cols = ['Date', 'Teacher_ID', 'Name', 'Time_In']
                missing_cols = [col for col in required_cols if col not in df.columns]
                if missing_cols:
                    errors.append(f"Missing columns: {missing_cols}")
            
            if errors:
                return False, f"Validation failed: {'; '.join(errors)}", warnings
            else:
                return True, "Validation passed", warnings
                
        except Exception as e:
            return False, f"Error validating file: {str(e)}", []
    
    def _format_data_sheet(self, ws, columns: List[str]):
        """Format data sheet with headers"""
        # Format headers
        for col_num, header in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.border = self.border
            cell.alignment = self.center_alignment
        
        # Auto-adjust columns
        self._auto_adjust_columns(ws)
    
    def _auto_adjust_columns(self, ws):
        """Auto-adjust column widths"""
        try:
            for column in ws.columns:
                max_length = 0
                column_letter = None
                
                # Get column letter from first cell
                for cell in column:
                    if cell.column_letter:
                        column_letter = cell.column_letter
                        break
                
                if not column_letter:
                    continue
                
                # Calculate max length
                for cell in column:
                    try:
                        if cell.value and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                # Set column width
                adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
                ws.column_dimensions[column_letter].width = adjusted_width
        except Exception as e:
            # If auto-adjust fails, continue without it
            pass
    
    def _get_working_days_in_month(self, year: int, month: int) -> int:
        """Calculate working days in a month (excluding weekends)"""
        from calendar import monthrange
        
        _, days_in_month = monthrange(year, month)
        working_days = 0
        
        for day in range(1, days_in_month + 1):
            date_obj = date(year, month, day)
            if date_obj.weekday() < 5:  # Monday = 0, Friday = 4
                working_days += 1
        
        return working_days
    
    def _calculate_average_time(self, time_list: List[str]) -> str:
        """Calculate average time from list of time strings"""
        try:
            total_seconds = 0
            valid_times = 0
            
            for time_str in time_list:
                if pd.isna(time_str) or time_str == '' or time_str is None:
                    continue
                
                try:
                    # Handle different time formats
                    time_str = str(time_str).strip()
                    if ':' in time_str:
                        # Parse time string
                        if len(time_str.split(':')) == 3:
                            time_obj = datetime.strptime(time_str, '%H:%M:%S').time()
                        else:
                            time_obj = datetime.strptime(time_str, '%H:%M').time()
                        
                        total_seconds += time_obj.hour * 3600 + time_obj.minute * 60 + time_obj.second
                        valid_times += 1
                except Exception as e:
                    # Skip invalid time entries
                    continue
            
            if valid_times == 0:
                return "N/A"
            
            avg_seconds = total_seconds // valid_times
            hours = avg_seconds // 3600
            minutes = (avg_seconds % 3600) // 60
            
            return f"{hours:02d}:{minutes:02d}"
            
        except Exception:
            return "N/A"
    
    def _format_summary_sheet(self, ws):
        """Format summary sheet"""
        ws.append(['Attendance Summary Report'])
        ws.append(['Generated on:', datetime.now().strftime('%Y-%m-%d %H:%M:%S')])
        ws.append([''])
        
        # Format title
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:D1')
    
    def _format_daily_sheet(self, ws):
        """Format daily attendance sheet"""
        headers = ['Date', 'Teacher_ID', 'Name', 'Time_In', 'Status', 'Confidence']
        ws.append(headers)
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = self.header_font
            cell.fill = self.header_fill
    
    def _format_monthly_sheet(self, ws):
        """Format monthly report sheet"""
        headers = ['Teacher', 'Days Present', 'Total Days', 'Attendance %']
        ws.append(headers)
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = self.header_font
            cell.fill = self.header_fill
    
    def _format_stats_sheet(self, ws):
        """Format statistics sheet"""
        headers = ['Metric', 'Value', 'Description']
        ws.append(headers)
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = self.header_font
            cell.fill = self.header_fill
    
    def _create_attendance_summary(self, ws, data_df, start_date, end_date):
        """Create attendance summary in worksheet"""
        ws.append(['Attendance Summary'])
        ws.append([f'Period: {start_date} to {end_date}'])
        ws.append([''])
        
        # Summary statistics
        total_records = len(data_df)
        unique_teachers = data_df['Teacher_ID'].nunique()
        unique_dates = data_df['Date'].nunique()
        
        ws.append(['Total Records:', total_records])
        ws.append(['Unique Teachers:', unique_teachers])
        ws.append(['Days Covered:', unique_dates])
    
    def _create_daily_statistics(self, ws, data_df):
        """Create daily statistics"""
        daily_stats = data_df.groupby('Date').agg({
            'Teacher_ID': 'count',
            'Recognition_Confidence': 'mean'
        }).round(2)
        
        ws.append(['Date', 'Teachers Present', 'Avg Confidence'])
        
        for date, row in daily_stats.iterrows():
            ws.append([date, row['Teacher_ID'], row['Recognition_Confidence']])
    
    def _create_teacher_performance(self, ws, data_df):
        """Create teacher performance analysis"""
        try:
            teacher_stats = data_df.groupby(['Teacher_ID', 'Name']).agg({
                'Date': 'count',
                'Recognition_Confidence': 'mean'
            }).round(2)
            
            ws.append(['Teacher ID', 'Name', 'Days Present', 'Avg Confidence', 'Avg Time'])
            
            for (teacher_id, name), row in teacher_stats.iterrows():
                # Calculate average time separately
                teacher_times = data_df[data_df['Teacher_ID'] == teacher_id]['Time_In'].tolist()
                avg_time = self._calculate_average_time(teacher_times)
                
                ws.append([teacher_id, name, row['Date'], row['Recognition_Confidence'], avg_time])
                
        except Exception as e:
            # Add error handling
            ws.append(['Teacher ID', 'Name', 'Days Present', 'Avg Confidence', 'Avg Time'])
            ws.append(['Error loading performance data', str(e), '', '', ''])
    
    def _add_attendance_chart(self, ws, num_teachers):
        """Add attendance chart to worksheet"""
        try:
            # Only create chart if we have data and reasonable number of teachers
            if num_teachers == 0 or num_teachers > 20:
                return
            
            # Create a simple bar chart
            chart = BarChart()
            chart.title = "Attendance Overview"
            chart.y_axis.title = "Days Present"
            chart.x_axis.title = "Teachers"
            
            # Calculate actual data range
            data_start_row = 4  # Assuming headers are in row 3
            data_end_row = min(data_start_row + num_teachers - 1, 20)
            
            if data_end_row >= data_start_row:
                # Add data (Days Present column - column 4)
                data = Reference(ws, min_col=4, min_row=data_start_row, max_row=data_end_row, max_col=4)
                # Categories (Teacher names - column 2)
                categories = Reference(ws, min_col=2, min_row=data_start_row, max_row=data_end_row)
                
                chart.add_data(data, titles_from_data=False)
                chart.set_categories(categories)
                
                # Position chart to the right of the data
                ws.add_chart(chart, "J3")
            
        except Exception as e:
            # Chart creation is optional, don't fail the entire operation
            pass

    def get_excel_statistics(self) -> Dict[str, Any]:
        """Get statistics about Excel files"""
        try:
            stats = {
                'teachers_count': 0,
                'attendance_records': 0,
                'file_sizes': {},
                'last_modified': {},
                'data_range': {}
            }
            
            # Teachers statistics
            if os.path.exists(self.teachers_file):
                teachers_df = pd.read_excel(self.teachers_file)
                stats['teachers_count'] = len(teachers_df)
                stats['file_sizes']['teachers'] = os.path.getsize(self.teachers_file)
                stats['last_modified']['teachers'] = datetime.fromtimestamp(
                    os.path.getmtime(self.teachers_file)
                ).strftime('%Y-%m-%d %H:%M:%S')
            
            # Attendance statistics
            if os.path.exists(self.attendance_file):
                attendance_df = pd.read_excel(self.attendance_file)
                stats['attendance_records'] = len(attendance_df)
                stats['file_sizes']['attendance'] = os.path.getsize(self.attendance_file)
                stats['last_modified']['attendance'] = datetime.fromtimestamp(
                    os.path.getmtime(self.attendance_file)
                ).strftime('%Y-%m-%d %H:%M:%S')
                
                if not attendance_df.empty:
                    stats['data_range'] = {
                        'start_date': attendance_df['Date'].min(),
                        'end_date': attendance_df['Date'].max()
                    }
            
            return stats
            
        except Exception as e:
            st.error(f"Error getting Excel statistics: {str(e)}")
            return {}